################## - Gmail library
import os.path

from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

###################
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

workbook = load_workbook ('Excel\Email Data.xlsx')

for k in workbook.sheetnames:
    del_sheet = workbook[k]
    workbook.remove(del_sheet)
###################

# If modifying these scopes, delete the file token.json.
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']

user_id = 'me'
input_search = 'after:2022/06/12' ##search email after 12-Jun-2022-this one is on the gmail search engine

### Get messages ###
def get_msg (service, user_id, msg_id):
   service = get_service() 
   try:
      message_encode=service.users().messages().get(userId=user_id, id = msg_id).execute() ##get the encoeded message
      trunc_id = '***'+msg_id[12:]
      for header in reversed(message_encode["payload"]["headers"]):
        if header['name'] == 'Subject':
            o_subject = header['value']
        if header['name'] == 'From':
            o_sender = header['value']

   except HttpError as error: 
       print ('An error occured: {}'.format(error))
   return [o_sender, o_subject, trunc_id]

### List messages id ###
def get_id (service, user_id, search_string):
   service = get_service()
   try:
       search_id=service.users().messages().list(userId=user_id, q=search_string).execute()
       number_result = search_id["resultSizeEstimate"] ## a returned element from the dictionary of user().list()
       msg_id_list = []
       if number_result > 0:
           messages_id = search_id["messages"] ## access the "messages"
           for ids in messages_id:
               msg_id_list.append(ids['id'])

       #else:
           #print ("There was no result for that search, returning an empty string")
           #return ''

   except HttpError as error: 
       print ('An error occured: {}'.format(error))

   #print (msg_id_list) 
   return msg_id_list
    
### Function from google API ###
def get_service():

    creds = None
    # The file token.json stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())

        # Call the Gmail API
    service = build('gmail', 'v1', credentials=creds)

    return service

### Combine the previous Function ###
def generate_list(searching):
    service = get_service()
    i_msg_id_list = get_id (service, user_id, searching)
    msg_info_list = []
    for i_msg_id in i_msg_id_list:
    #i_msg_id = "18160e83d24ddc25"
        msg_info_list.append(get_msg (service, user_id, i_msg_id))

    final_msg_list = []
    if searching == "in:sent":
        final_msg_list = msg_info_list
    else:
        for j in range (len(msg_info_list)):
            if 'Re: ' not in msg_info_list[j][1]:
                final_msg_list.append (msg_info_list[j])

    return final_msg_list

def generate_sheet (searching):
    the_list = generate_list(searching)
    searching = ":"+searching
    indi, key_word = searching.rsplit(':', 1)

    if key_word == '':
        key_word = "BLANK_named"
    elif len(key_word) > 31:
        key_word = key_word[:28]+"..."
    elif key_word.isalnum() == 0:
        for char in key_word:
            if char in ":\/?*[]":
                key_word = key_word.replace (char," ")
    else:    
        key_word = key_word.capitalize()

    sheet = workbook.create_sheet(key_word)

    headings = ["From", "Subject", "ID (Technical)"]
    heading_cell = ['A1', 'B1', 'C1']
    sheet.append (headings)
    for email_data in the_list:
        sheet.append(email_data)
    
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 80
    sheet.column_dimensions['C'].width = 13
    for row in sheet.iter_rows():  
        for cell in row:      
            cell.alignment = Alignment(wrap_text=True)
    for cell in heading_cell:      
        sheet[cell].alignment = Alignment(horizontal='center') 
    


    workbook.save ('Excel\Email Data.xlsx')
    workbook.close()


cat_primary = generate_sheet("category:primary")
cat_social = generate_sheet("category:social")
cat_promo = generate_sheet("category:promotions")
unread = generate_sheet("label:unread")
starred = generate_sheet("is:starred")
sent = generate_sheet("in:sent")

look_for = generate_sheet(input_search)







